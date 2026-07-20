import argparse
import os
import sys

import pandas as pd

SYSTEMS = ["microbench_nomad", "microbench_tpp"]

TEST_FILES = [
    "zipfan_hottest_10G.read.log.csv",
    "zipfan_hottest_13.5G.read.log.csv",
    "zipfan_hottest_27G.read.log.csv",
    "zipfan_hottest_10G.write.log.csv",
    "zipfan_hottest_13.5G.write.log.csv",
    "zipfan_hottest_27G.write.log.csv",
]

RAW_COLUMNS = ["Bandwidth(MB/s)", "milliseconds", "total tick", "work type"]

# counters that are start_/end_ pairs but should show the raw end_ value
# instead of the end-start diff
RAW_INSTEAD_OF_DIFF = {"shadow_page_pair"}

# categories whose rows are raw values, so they don't get the "Δ " prefix
RAW_CATEGORIES = {"성능"}


# ---- row definition helpers -------------------------------------------
# A row shows a value from "nomad_col"/"tpp_col" depending on which system's
# CSV is being processed. If the column for that system is None, or missing
# from the CSV, the row's cells are left blank.

def sep(category):
    return {"kind": "sep", "label": f"[{category}]", "category": category}


def row(label, nomad_col=None, tpp_col=None):
    return {"kind": "row", "label": label, "nomad_col": nomad_col, "tpp_col": tpp_col}


def both(label, col):
    return row(label, nomad_col=col, tpp_col=col)


def ratio(label, nomad_cols=None, tpp_cols=None):
    return {"kind": "ratio", "label": label, "nomad_cols": nomad_cols, "tpp_cols": tpp_cols}


def total(label, labels):
    return {"kind": "sum", "label": label, "labels": labels}


KEY_ROWS = [
    sep("마이그레이션 횟수"),
    row("승격 성공 횟수", nomad_col="success_nr", tpp_col="pgpromote_anon"),
    both("강등 성공 횟수", "pgdemote_anon"),
    total("총합", ["승격 성공 횟수", "강등 성공 횟수"]),

    sep("마이그레이션 오버헤드 감소"),
    row("transactional_migration_success_num", nomad_col="transactional_migration_success_num"),
    row("shadow_demote_num", nomad_col="shadow_demote_num"),
    row("pgdemote_anon", nomad_col="pgdemote_anon"),
    ratio("shadow_demote_num / pgdemote_anon", nomad_cols=("shadow_demote_num", "pgdemote_anon")),
    row("NUMA_HINT_FAULTS 또는 pgpromote_candidate_nomad",
        tpp_col=("numa_hint_faults", "numa_hint_faults_local"),
        nomad_col=("pgpromote_candidate_nomad", "numa_hint_faults_local")),
    row("pgpromote_anon 또는 success_nr", tpp_col="pgpromote_anon", nomad_col="success_nr"),
    ratio("NUMA_HINT_FAULTS / pgpromote_anon 또는 pgpromote_candidate_nomad / success_nr",
          tpp_cols=(("numa_hint_faults", "numa_hint_faults_local"), "pgpromote_anon"),
          nomad_cols=(("pgpromote_candidate_nomad", "numa_hint_faults_local"), "success_nr")),

    sep("write 워크로드에서"),
    row("write_protect_break_num", nomad_col="write_protect_break_num"),
    row("transactional_migration_fail_num", nomad_col="transactional_migration_fail_num"),

    sep("SSP 모티베이션"),
    row("write_protect_break_num", nomad_col="write_protect_break_num"),
    row("shadow_page_pair", nomad_col="shadow_page_pair"),
    row("batch_free_num", nomad_col="batch_free_num"),

    sep("성능"),
    both("Bandwidth(MB/s)", "Bandwidth(MB/s)"),
]

ALL_ROWS = [
    sep("마이그레이션 횟수: 바닐라 승격"),
    both("pgpromote_success", "pgpromote_success"),
    both("pgpromote_candidate", "pgpromote_candidate"),
    both("pgpromote_tried", "pgpromote_tried"),
    both("pgpromote_file", "pgpromote_file"),
    both("pgpromote_anon", "pgpromote_anon"),

    sep("마이그레이션 횟수: 강등"),
    both("pgdemote_kswapd", "pgdemote_kswapd"),
    both("pgdemote_direct", "pgdemote_direct"),
    both("pgdemote_khugepaged", "pgdemote_khugepaged"),
    both("pgdemote_file", "pgdemote_file"),
    both("pgdemote_anon", "pgdemote_anon"),

    sep("승격 후보"),
    both("pgpromote_candidate_nomad", "pgpromote_candidate_nomad"),
    both("pgpromote_candidate_demoted_nomad", "pgpromote_candidate_demoted_nomad"),
    both("pgpromote_candidate_anon_nomad", "pgpromote_candidate_anon_nomad"),
    both("pgpromote_candidate_file_nomad", "pgpromote_candidate_file_nomad"),
    both("numa_hint_faults", "numa_hint_faults"),
    both("numa_hint_faults_local", "numa_hint_faults_local"),

    sep("실패 사유"),
    both("pgmigrate_dst_node_full_fail", "pgmigrate_dst_node_full_fail"),
    both("pgmigrate_numa_isolate_fail", "pgmigrate_numa_isolate_fail"),
    both("pgmigrate_refcount_fail", "pgmigrate_refcount_fail"),
    both("pgmigrate_nomem_fail", "pgmigrate_nomem_fail"),
    both("pgmigrate_nomem_fail_promote", "pgmigrate_nomem_fail_promote"),
    both("pgmigrate_nomem_fail_demote", "pgmigrate_nomem_fail_demote"),

    sep("커널모듈: 승격"),
    row("try_to_promote_nr", nomad_col="try_to_promote_nr"),
    row("retreated_page_nr", nomad_col="retreated_page_nr"),
    row("transactional_migration_success_num", nomad_col="transactional_migration_success_num"),
    row("transactional_migration_fail_num", nomad_col="transactional_migration_fail_num"),

    sep("커널모듈: 셰도우 페이지"),
    row("shadow_link_num", nomad_col="shadow_link_num"),
    row("shadow_demote_num", nomad_col="shadow_demote_num"),
    row("write_protect_break_num", nomad_col="write_protect_break_num"),
    row("batch_free_num", nomad_col="batch_free_num"),
    row("shadow_page_pair", nomad_col="shadow_page_pair"),

    sep("성능"),
    both("milliseconds", "milliseconds"),
    both("total tick", "total tick"),
    both("work type", "work type"),
    both("Bandwidth(MB/s)", "Bandwidth(MB/s)"),
]


# ---- value computation ---------------------------------------------------

def compute_values(df):
    """counter name -> list of per-round values (end-start diff, or raw)."""
    values = {}
    for col in df.columns:
        if col.startswith("start_"):
            name = col[len("start_"):]
            end_col = "end_" + name
            if end_col in df.columns:
                if name in RAW_INSTEAD_OF_DIFF:
                    values[name] = df[end_col].tolist()
                else:
                    values[name] = (df[end_col] - df[col]).tolist()
    for col in RAW_COLUMNS:
        if col in df.columns:
            values[col] = df[col].tolist()
    return values


def lookup_row_value(values, col, n):
    """col is either a counter name, or (a, b) meaning "a - b"."""
    if col is None:
        return [None] * n
    if isinstance(col, tuple):
        a_name, b_name = col
        a_vals = values.get(a_name)
        b_vals = values.get(b_name)
        if a_vals is None or b_vals is None:
            return [None] * n
        return [a - b for a, b in zip(a_vals, b_vals)]
    return values.get(col, [None] * n)


def compute_ratio(values, cols, n):
    if cols is None:
        return [None] * n
    num_col, den_col = cols
    nums = lookup_row_value(values, num_col, n)
    dens = lookup_row_value(values, den_col, n)
    return [a / b if (a is not None and b not in (0, None)) else None for a, b in zip(nums, dens)]


def compute_sum(computed, labels, n):
    parts = [computed[label] for label in labels]
    return [sum(vs) if all(v is not None for v in vs) else None for vs in zip(*parts)]


def build_sheet(df, system, row_defs):
    values = compute_values(df)
    n = len(df)
    computed = {}
    rows = []
    sep_positions = []
    current_category = None

    for row_def in row_defs:
        kind = row_def["kind"]
        label = row_def["label"]

        if kind == "sep":
            current_category = row_def["category"]
            vals = [None] * n
            sep_positions.append(len(rows))
            display_label = label
        else:
            if kind == "row":
                col = row_def["nomad_col"] if system == "microbench_nomad" else row_def["tpp_col"]
                vals = lookup_row_value(values, col, n)
            elif kind == "ratio":
                cols = row_def["nomad_cols"] if system == "microbench_nomad" else row_def["tpp_cols"]
                vals = compute_ratio(values, cols, n)
            elif kind == "sum":
                vals = compute_sum(computed, row_def["labels"], n)

            if label in RAW_INSTEAD_OF_DIFF or current_category in RAW_CATEGORIES:
                display_label = label
            else:
                display_label = f"Δ {label}"

        computed[label] = vals
        rows.append((display_label, vals))

    columns = [f"r{i}" for i in range(n)]
    sheet_df = pd.DataFrame([vals for _, vals in rows], index=[label for label, _ in rows], columns=columns)
    return sheet_df, sep_positions


def apply_separator_style(worksheet, sep_positions, n_cols):
    """sep_positions: 0-based row positions (within the data rows) to shade gray across A:<last round col>."""
    from openpyxl.styles import PatternFill

    fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    last_col = 1 + n_cols  # +1 for the label column
    for pos in sep_positions:
        excel_row = pos + 2  # +1 for header row, +1 for 1-indexing
        for col in range(1, last_col + 1):
            worksheet.cell(row=excel_row, column=col).fill = fill


def apply_label_column_style(worksheet, n_rows):
    """Column A (counter labels): left-aligned, not bold."""
    from openpyxl.styles import Alignment, Font

    for excel_row in range(1, n_rows + 2):  # +1 for header row, +1 for 1-indexing
        cell = worksheet.cell(row=excel_row, column=1)
        cell.alignment = Alignment(horizontal="left")
        cell.font = Font(bold=False)


# ---- main -----------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="microbench CSV를 카운터 요약 xlsx로 가공")
    parser.add_argument("--base-dir", default="src/post_processing/tmp",
                         help="microbench_nomad/microbench_tpp 폴더가 들어있는 경로")
    args = parser.parse_args()

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("openpyxl이 설치되어 있지 않습니다. 'pip install openpyxl' 후 다시 실행하세요.")
        sys.exit(1)

    for system in SYSTEMS:
        in_dir = os.path.join(args.base_dir, system)
        out_dir = os.path.join(in_dir, "summary")
        os.makedirs(out_dir, exist_ok=True)

        for test_file in TEST_FILES:
            in_path = os.path.join(in_dir, test_file)
            if not os.path.exists(in_path):
                print(f"{in_path} 없음, 건너뜀")
                continue

            df = pd.read_csv(in_path, index_col=0)
            key_sheet, key_seps = build_sheet(df, system, KEY_ROWS)
            all_sheet, all_seps = build_sheet(df, system, ALL_ROWS)

            out_name = test_file.replace(".csv", "") + ".summary.xlsx"
            out_path = os.path.join(out_dir, out_name)
            with pd.ExcelWriter(out_path) as writer:
                key_sheet.to_excel(writer, sheet_name="주요 카운터")
                all_sheet.to_excel(writer, sheet_name="전체")
                apply_separator_style(writer.sheets["주요 카운터"], key_seps, key_sheet.shape[1])
                apply_label_column_style(writer.sheets["주요 카운터"], key_sheet.shape[0])
                apply_separator_style(writer.sheets["전체"], all_seps, all_sheet.shape[1])
                apply_label_column_style(writer.sheets["전체"], all_sheet.shape[0])
                writer.sheets["주요 카운터"].column_dimensions["A"].width = 40
                writer.sheets["전체"].column_dimensions["A"].width = 40
            print(f"saved {out_path}")


if __name__ == "__main__":
    main()
